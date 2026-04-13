"""
exporter.py — Génération Excel 6 onglets avec mise en forme et hyperliens dashboard.
"""

import logging
import os
import shutil
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY
from openpyxl.worksheet.hyperlink import Hyperlink

import database as db

logger = logging.getLogger(__name__)

DASHBOARD_URL = os.getenv("DASHBOARD_URL", "http://192.168.137.1:5000")
EXCEL_PATH = os.getenv("EXCEL_PATH", "presences.xlsx")
SHARED_FOLDER = os.getenv("SHARED_FOLDER", "")  # Chemin réseau ou Google Drive

# ─── Couleurs ────────────────────────────────────────────────────────────────
COLOR_HEADER_BG   = "1A237E"  # Bleu marine profond
COLOR_HEADER_FG   = "FFFFFF"
COLOR_COMPLET     = "C8E6C9"  # Vert clair
COLOR_INCOMPLET   = "FFE0B2"  # Orange clair
COLOR_ABSENT      = "FFCDD2"  # Rouge clair
COLOR_COMPLET_TXT = "1B5E20"
COLOR_INCOMPLET_TXT = "E65100"
COLOR_ABSENT_TXT  = "B71C1C"
COLOR_ROW_ALT     = "F5F5F5"
COLOR_BORDER      = "BDBDBD"

FILLS = {
    "Complet":   PatternFill("solid", fgColor=COLOR_COMPLET),
    "Incomplet": PatternFill("solid", fgColor=COLOR_INCOMPLET),
    "Absent":    PatternFill("solid", fgColor=COLOR_ABSENT),
}
FONTS_STATUS = {
    "Complet":   Font(color=COLOR_COMPLET_TXT, bold=True, size=9),
    "Incomplet": Font(color=COLOR_INCOMPLET_TXT, bold=True, size=9),
    "Absent":    Font(color=COLOR_ABSENT_TXT, bold=True, size=9),
}

THIN = Side(border_style="thin", color=COLOR_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_fill():
    return PatternFill("solid", fgColor=COLOR_HEADER_BG)


def _header_font():
    return Font(color=COLOR_HEADER_FG, bold=True, size=10)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_header(cell, text):
    cell.value = text
    cell.fill = _header_fill()
    cell.font = _header_font()
    cell.alignment = _center()
    cell.border = BORDER


def _format_dur(minutes):
    """Formate les minutes en H:MM."""
    if minutes is None:
        return ""
    h = minutes // 60
    m = minutes % 60
    return f"{h}:{m:02d}"


def _add_hyperlink_cell(ws, row, col, display_text, url):
    """Ajoute une cellule avec hyperlien."""
    cell = ws.cell(row=row, column=col)
    cell.value = display_text
    cell.hyperlink = url
    cell.font = Font(color="1565C0", underline="single", bold=True, size=9)
    cell.alignment = _center()
    cell.border = BORDER
    return cell


# ─── Onglet Jour ─────────────────────────────────────────────────────────────

def _build_onglet_jour(ws, date_str: str):
    ws.title = "Jour"
    ws.sheet_view.showGridLines = False

    # Titre
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"KAROKA — Présences du {date_str}"
    title_cell.font = Font(color=COLOR_HEADER_FG, bold=True, size=13)
    title_cell.fill = _header_fill()
    title_cell.alignment = _center()

    headers = [
        "PRNO", "Nom Prénom", "Date",
        "Arr. Matin", "Dép. Matin", "Arr. AM", "Dép. AM",
        "Dur. Matin", "Dur. AM", "Dur. Totale", "Statut"
    ]
    for col, h in enumerate(headers, 1):
        _apply_header(ws.cell(row=2, column=col), h)

    resume = db.get_resume_jour(date_str)
    for i, r in enumerate(resume):
        row = i + 3
        fill = PatternFill("solid", fgColor=COLOR_ROW_ALT) if i % 2 == 0 else PatternFill()
        values = [
            r["prno"],
            r["nom_prenom"],
            date_str,
            r["arr_mat"] or "",
            r["dep_mat"] or "",
            r["arr_apm"] or "",
            r["dep_apm"] or "",
            _format_dur(r["dur_mat"]),
            _format_dur(r["dur_apm"]),
            _format_dur(r["dur_tot"]),
            r["statut"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = _center()
            cell.border = BORDER
            if col == 11:  # Statut
                cell.fill = FILLS.get(r["statut"], PatternFill())
                cell.font = FONTS_STATUS.get(r["statut"], Font(size=9))
            else:
                cell.fill = fill
                cell.font = Font(size=9)

    # Largeurs
    col_widths = [10, 22, 12, 10, 10, 10, 10, 10, 10, 12, 12]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "A3"


# ─── Onglet Période ──────────────────────────────────────────────────────────

def _build_onglet_periode(ws, titre: str, date_debut: str, date_fin: str):
    ws.title = titre
    ws.sheet_view.showGridLines = False

    # Dates (jours ouvrables)
    start = date.fromisoformat(date_debut)
    end = date.fromisoformat(date_fin)
    all_dates = []
    d = start
    while d <= end:
        if d.weekday() < 5:
            all_dates.append(d.isoformat())
        d += timedelta(days=1)

    # Titre
    total_cols = 2 + len(all_dates)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"KAROKA — {titre} ({date_debut} → {date_fin})"
    title_cell.font = Font(color=COLOR_HEADER_FG, bold=True, size=13)
    title_cell.fill = _header_fill()
    title_cell.alignment = _center()

    # Headers fixes
    _apply_header(ws.cell(row=2, column=1), "PRNO")
    _apply_header(ws.cell(row=2, column=2), "Nom Prénom")

    # Headers dates avec hyperlien
    for col_offset, ds in enumerate(all_dates):
        col = col_offset + 3
        dt = date.fromisoformat(ds)
        label = f"{dt.strftime('%a %d/%m')}"
        url = f"{DASHBOARD_URL}/?date={ds}"
        _add_hyperlink_cell(ws, row=2, col=col, display_text=label, url=url)
        ws.cell(row=2, column=col).fill = _header_fill()
        ws.cell(row=2, column=col).font = Font(
            color=COLOR_HEADER_FG, bold=True, size=9, underline="single"
        )

    # Données
    resume = db.get_resume_periode(date_debut, date_fin)
    for i, emp in enumerate(resume):
        row = i + 3
        alt_fill = PatternFill("solid", fgColor=COLOR_ROW_ALT) if i % 2 == 0 else PatternFill()

        cell_prno = ws.cell(row=row, column=1, value=emp["prno"])
        cell_prno.alignment = _center()
        cell_prno.border = BORDER
        cell_prno.fill = alt_fill
        cell_prno.font = Font(size=9, bold=True)

        cell_nom = ws.cell(row=row, column=2, value=emp["nom_prenom"])
        cell_nom.alignment = Alignment(horizontal="left", vertical="center")
        cell_nom.border = BORDER
        cell_nom.fill = alt_fill
        cell_nom.font = Font(size=9)

        for col_offset, ds in enumerate(all_dates):
            col = col_offset + 3
            statut = emp["dates"].get(ds, "Absent")
            cell = ws.cell(row=row, column=col, value=statut)
            cell.fill = FILLS.get(statut, PatternFill())
            cell.font = FONTS_STATUS.get(statut, Font(size=8))
            cell.alignment = _center()
            cell.border = BORDER
            # Hyperlien sur la cellule de statut
            url = f"{DASHBOARD_URL}/?date={ds}"
            cell.hyperlink = url

    # Largeurs
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    for col_offset in range(len(all_dates)):
        ws.column_dimensions[get_column_letter(col_offset + 3)].width = 11
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 36
    ws.freeze_panes = "C3"


# ─── Génération principale ────────────────────────────────────────────────────

def generate_excel(target_date: str = None):
    """Génère le fichier Excel complet."""
    if target_date is None:
        from zoneinfo import ZoneInfo
        tz = ZoneInfo(os.getenv("TIMEZONE", "Indian/Antananarivo"))
        from datetime import datetime
        target_date = datetime.now(tz).date().isoformat()

    today = date.fromisoformat(target_date)

    # Calculs des périodes
    lundi = today - timedelta(days=today.weekday())
    dimanche = lundi + timedelta(days=6)
    premier_mois = today.replace(day=1)
    if today.month <= 3:
        premier_trim = today.replace(month=1, day=1)
    elif today.month <= 6:
        premier_trim = today.replace(month=4, day=1)
    elif today.month <= 9:
        premier_trim = today.replace(month=7, day=1)
    else:
        premier_trim = today.replace(month=10, day=1)

    if today.month <= 6:
        premier_sem = today.replace(month=1, day=1)
        fin_sem = today.replace(month=6, day=30)
    else:
        premier_sem = today.replace(month=7, day=1)
        fin_sem = today.replace(month=12, day=31)

    premier_an = today.replace(month=1, day=1)
    fin_an = today.replace(month=12, day=31)

    # Fin de trimestre
    mois_fin_trim = ((premier_trim.month - 1) // 3 + 1) * 3
    import calendar
    dernier_jour = calendar.monthrange(today.year, mois_fin_trim)[1]
    fin_trim = today.replace(month=mois_fin_trim, day=dernier_jour)

    # Fin de mois
    fin_mois_day = calendar.monthrange(today.year, today.month)[1]
    fin_mois = today.replace(day=fin_mois_day)

    wb = Workbook()
    # Supprimer la feuille par défaut
    wb.remove(wb.active)

    # Onglet Jour
    ws_jour = wb.create_sheet("Jour")
    _build_onglet_jour(ws_jour, target_date)

    # Onglets périodiques
    periodes = [
        ("Semaine",    lundi.isoformat(),       dimanche.isoformat()),
        ("Mois",       premier_mois.isoformat(), fin_mois.isoformat()),
        ("Trimestre",  premier_trim.isoformat(), fin_trim.isoformat()),
        ("Semestre",   premier_sem.isoformat(),  fin_sem.isoformat()),
        ("Année",      premier_an.isoformat(),   fin_an.isoformat()),
    ]
    for titre, d_debut, d_fin in periodes:
        ws = wb.create_sheet(titre)
        _build_onglet_periode(ws, titre, d_debut, d_fin)

    # Sauvegarder
    wb.save(EXCEL_PATH)
    logger.info("Excel généré : %s", EXCEL_PATH)

    # Copie sur dossier partagé
    if SHARED_FOLDER and Path(SHARED_FOLDER).exists():
        dest = Path(SHARED_FOLDER) / "presences.xlsx"
        shutil.copy2(EXCEL_PATH, dest)
        logger.info("Excel copié sur dossier partagé : %s", dest)

    return EXCEL_PATH


if __name__ == "__main__":
    import sys
    from dotenv import load_dotenv
    load_dotenv()
    db.init_db()
    date_arg = sys.argv[1] if len(sys.argv) > 1 else None
    path = generate_excel(date_arg)
    print(f"Excel généré : {path}")
